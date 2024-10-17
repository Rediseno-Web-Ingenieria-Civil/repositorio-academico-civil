import argparse
import json
import pandas as pd
from openpyxl import load_workbook

file_name = "src/data/DIC-UCHILE.xlsx"  # path to file + file name

# Parse command line arguments
parser = argparse.ArgumentParser()
parser.add_argument("--file", help="Path to the Excel file", default=file_name)
args = parser.parse_args()

file_name = args.file

# Load the workbook and select the sheet
wb = load_workbook(filename=file_name)

allowed_sheets = ["ACADÉMICOS", "WEB"]

meses = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]

mes_to_int = {
    "Enero": 1,
    "Febrero": 2,
    "Marzo": 3,
    "Abril": 4,
    "Mayo": 5,
    "Junio": 6,
    "Julio": 7,
    "Agosto": 8,
    "Septiembre": 9,
    "Octubre": 10,
    "Noviembre": 11,
    "Diciembre": 12,
}

mes_inicial = "Enero"  # @param ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
ano_inicial = 2022  # @param {"type":"slider","min":2020,"max":2040,"step":1}

mes_final = "Septiembre"  # @param ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
ano_final = 2024  # @param {"type":"slider","min":2020,"max":2040,"step":1}


def check_sheet_name(sheet_name: str) -> bool:
    if sheet_name in allowed_sheets:
        return True

    try:
        mes, ano = sheet_name.split(" ")
        if mes in meses:
            if int(ano) >= ano_inicial:
                return True
            else:
                print(f"{mes} {ano}: El año debe ser mayor o igual a {ano_inicial}")
        else:
            print(
                f"{mes} {ano}: El mes no está en la lista de meses, revisar que esté bien escrito"
            )
    except Exception as e:
        print(
            f"Hoja {sheet_name} no cumple formato de nombre 'Mes año' (ej: Enero 2022)"
        )
        print(f"\tError: {e}")
    return False


# iterate over all sheets
names_failed = False
for sheet in wb.sheetnames:
    if not check_sheet_name(sheet):
        names_failed = True

if names_failed:
    print("Error: Al menos una hoja no cumple con el formato de nombre 'Mes año'")
    exit()

areas = set()
rhma = "Recursos Hídricos y Medio Ambiente"
ecg = "Estructuras y Geotecnia"
tran = "Transporte"
area_map = {
    "Recursos Hídricos y Medio Ambiente": rhma,
    "Recursos hídricos y medio ambiente": rhma,
    "RHMA": rhma,
    "Recursos Hídricos y Medio ambiente": rhma,
    "ECG": ecg,
    "Estructura y Geotécnia": ecg,
    "Estructuras y Geotecnia": ecg,
    "Estructura y Geotecnia": ecg,
    "Transporte": tran,
}
repositorio_data = {}
# for ano in range(ano_inicial, ano_final+1):
for ano in range(ano_final, ano_inicial - 1, -1):
    ano_data = {}
    # for mes in meses:
    for mes in meses[::-1]:
        if ano == ano_inicial and mes not in meses[meses.index(mes_inicial) :]:
            continue
        if ano == ano_final and mes not in meses[: meses.index(mes_final) + 1]:
            continue
        sheet_name = f"{mes} {ano}"
        if not sheet_name in wb.sheetnames:
            print(f"Error: No se encontró la hoja {sheet_name}")
            continue
        sheet = wb[sheet_name]

        # Extract data from the sheet
        mes_data = []
        links_publicacion = {}
        links_publicado_en = {}
        links_doi = {}
        # autores_dic = set()

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            mes_data.append(row)
            if row[3] and (hyperlink := sheet.cell(row=row_idx, column=4).hyperlink):
                links_publicacion[row[3]] = hyperlink.target

            if row[4] and (hyperlink := sheet.cell(row=row_idx, column=5).hyperlink):
                links_publicado_en[row[4]] = hyperlink.target

            if row[5] and (hyperlink := sheet.cell(row=row_idx, column=6).hyperlink):
                links_doi[row[5]] = hyperlink.target

            # Check if the cell in column 2 has a fill color
            # No funcionan, solo retornan '00000000'
            # if row[2]: print(row[2], sheet.cell(row=row_idx, column=2).fill.start_color.index)
            # if row[2]: print(row[2], sheet.cell(row=row_idx, column=2).fill.__dict__.items(), "\n")
            # if sheet.cell(row=row_idx, column=2).fill.start_color.index != "00000000":
            #     autores_dic.add(row[2])

        # Convert to DataFrame
        df = pd.DataFrame(mes_data).dropna(axis=0, how="all").dropna(axis=1, how="all")

        # Expected columns
        expected_cols = [
            "N°",
            "AREA",
            "AUTORES",
            "NOMBRE DE LA PUBLICACIÓN",
            "PUBLICADO EN",
            "DOI",
            "FECHA",
            "WOS",
            "SCOPUS",
            "ORCID",
            "Veces Citado",
        ]

        # Iterate through rows until the headers match the expected columns
        for i in range(len(df)):
            header = df.iloc[i].to_list()
            if header == expected_cols:
                df = df.iloc[i + 1 :]  # Update DataFrame to start from the next row
                df.columns = expected_cols  # Set the correct headers
                break

        # Forward fill na values in column "NOMBRE DE LA PUBLICACIÓN"
        df["NOMBRE DE LA PUBLICACIÓN"] = df["NOMBRE DE LA PUBLICACIÓN"].ffill()

        # Group by "NOMBRE DE LA PUBLICACIÓN" and combine other columns as lists
        grouped_df = df.groupby("NOMBRE DE LA PUBLICACIÓN").agg(
            lambda x: x.dropna().tolist()
        )

        # if element is a list with one element, extract the element
        grouped_df = grouped_df.map(lambda x: x[0] if len(x) == 1 else x)

        # Reset DataFrame index
        grouped_df.reset_index(inplace=True)

        # add columns for hyperlinks
        grouped_df["LINK PUBLICACION"] = grouped_df["NOMBRE DE LA PUBLICACIÓN"].map(
            links_publicacion
        )
        grouped_df["LINK PUBLICADO EN"] = grouped_df["PUBLICADO EN"].map(
            links_publicado_en
        )
        grouped_df["LINK DOI"] = grouped_df["DOI"].map(links_doi)

        # Add columns for year and month
        grouped_df["MES"] = mes
        grouped_df["ANO"] = ano

        # Drop column FECHA
        # grouped_df.drop(columns=["FECHA"], inplace=True)
        grouped_df["FECHA"] = f"{ano}-{mes_to_int[mes]:02d}-01"

        # Map areas
        grouped_df["AREA"] = grouped_df["AREA"].map(
            lambda x: area_map[x] if not isinstance(x, list) and x in area_map else x
        )

        # Append to ano_data
        ano_data[f"{mes}"] = grouped_df.to_dict(orient="records")

        area_df = pd.DataFrame()
        area_df["AREA"] = grouped_df["AREA"].map(
            lambda x: tuple(x) if isinstance(x, list) else x
        )
        # Add areas to set
        areas.update(area_df["AREA"].unique())

    # Append to repositorio_data
    # + " " to avoid javascript sorting the keys
    repositorio_data[f"{ano}" + " "] = ano_data

print("Áreas encontradas:", areas)
# Save areas to a JSON file
areas_data = list(areas)
with open("src/data/areas.json", "w") as f:
    json.dump(areas_data, f, ensure_ascii=False, indent=4)
with open("src/data/repositorio_academico.json", "w") as f:
    json.dump(repositorio_data, f, ensure_ascii=False, indent=4)
